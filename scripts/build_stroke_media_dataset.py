#!/usr/bin/env python3
"""Build a paired stroke media dataset from the exported audit workbook."""

from __future__ import annotations

import argparse
import concurrent.futures
import csv
import hashlib
import json
import mimetypes
import os
import re
import shutil
import sys
import tempfile
import time
import urllib.error
import urllib.parse
import urllib.request
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - environment guard
    raise SystemExit(
        "Missing dependency: openpyxl. Install it or run with a Python "
        "environment that already has openpyxl."
    ) from exc


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE_NAME = (
    "\u8111\u5352\u4e2d\u6570\u636e\u91c7\u96c6-\u5ba1\u6838\u5bfc\u51fa-20260119.xlsx"
)
DEFAULT_SOURCE = PROJECT_ROOT / DEFAULT_SOURCE_NAME
DEFAULT_OUTPUT = PROJECT_ROOT / "datasets" / "stroke_media_dataset_20260119"

URL_RE = re.compile(r"https?://[^\s,，;；]+", re.IGNORECASE)

MEDIA_COLUMN_ROLES = {
    "\u8f85\u52a9\u68c0\u67e5\u5f71\u50cf": "auxiliary_exam_image",
    "\u75c5\u5386": "medical_record",
    "\u6b63\u9762\uff08URL\u5730\u5740\uff09": "front",
    "\u5fae\u7b11\uff08URL\u5730\u5740\uff09": "smile",
    "\u76b1\u989d\uff08URL\u5730\u5740\uff09": "forehead_wrinkle",
    "\u76b1\u7709\uff08URL\u5730\u5740\uff09": "frown",
    "\u95ed\u773c\uff08URL\u5730\u5740\uff09": "eyes_closed",
    "\u793a\u9f7f\uff08URL\u5730\u5740\uff09": "teeth",
    "\u4fa7\u89c6-\u5de6\uff08URL\u5730\u5740\uff09": "left_profile",
    "\u4fa7\u89c6-\u53f3\uff08URL\u5730\u5740\uff09": "right_profile",
    "\u820c\u50cf-\u9762\uff08URL\u5730\u5740\uff09": "tongue_surface",
    "\u820c\u50cf-\u5e95\uff08URL\u5730\u5740\uff09": "tongue_bottom",
    "\u97f3\u89c6\u9891\uff08URL\u5730\u5740\uff09": "audio_video",
    "\u6b63\u8138\u8f6e\u5ed3": "front_contour",
    "\u773c\u7403\u53f3\u770b": "eyes_right",
    "\u5fae\u7b11\u793a\u9f7f": "smile_teeth",
    "\u820c\u9762\u8f6e\u5ed3": "tongue_surface_contour",
    "\u820c\u6839\u8f6e\u5ed3": "tongue_root_contour",
    "\u89c6\u9891\u5730\u5740": "video",
    "\u8bca\u65ad\u62a5\u544a": "diagnostic_report",
}

IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".gif", ".tif", ".tiff"}
VIDEO_EXTS = {".mp4", ".mov", ".avi", ".mkv", ".webm", ".m4v", ".3gp"}

BASE_RECORD_FIELDS = [
    "record_id",
    "source_excel_row",
    "source_sheet",
    "source_file",
]

MEDIA_MANIFEST_FIELDS = [
    "media_id",
    "record_id",
    "source_excel_row",
    "source_sheet",
    "patient_id",
    "collection_no",
    "field_name",
    "media_role",
    "media_index_in_field",
    "media_type",
    "url",
    "local_path",
    "filename",
    "download_status",
    "http_status",
    "content_type",
    "bytes",
    "sha256",
    "error",
]


@dataclass(frozen=True)
class MediaItem:
    media_id: str
    record_id: str
    source_excel_row: int
    source_sheet: str
    patient_id: str
    collection_no: str
    field_name: str
    media_role: str
    media_index_in_field: int
    media_type: str
    url: str
    local_path: Path
    filename: str


def normalized_scalar(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def stable_id_part(value: Any, fallback: str) -> str:
    text = normalized_scalar(value)
    if not text:
        text = fallback
    text = re.sub(r"\.0$", "", text)
    text = re.sub(r"[^A-Za-z0-9_-]+", "-", text).strip("-")
    return text or fallback


def first_available(row: dict[str, Any], names: list[str], fallback: str) -> str:
    for name in names:
        value = normalized_scalar(row.get(name))
        if value:
            return stable_id_part(value, fallback)
    return fallback


def extract_urls(value: Any) -> list[str]:
    if value is None:
        return []
    urls = URL_RE.findall(str(value))
    return [url.rstrip(").]}>\"'") for url in urls]


def media_type_from_url(url: str) -> str:
    parsed = urllib.parse.urlsplit(url)
    ext = Path(parsed.path).suffix.lower()
    if ext in IMAGE_EXTS:
        return "image"
    if ext in VIDEO_EXTS:
        return "video"
    guessed, _ = mimetypes.guess_type(parsed.path)
    if guessed:
        if guessed.startswith("image/"):
            return "image"
        if guessed.startswith("video/"):
            return "video"
    return "unknown"


def extension_from_url(url: str, media_type: str) -> str:
    ext = Path(urllib.parse.urlsplit(url).path).suffix.lower()
    if ext:
        return ext
    if media_type == "image":
        return ".jpg"
    if media_type == "video":
        return ".mp4"
    return ".bin"


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def sha256_path(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def prepare_dirs(output: Path) -> None:
    for relative in [
        "metadata",
        "media/images",
        "media/videos",
        "media/unknown",
        "logs",
    ]:
        (output / relative).mkdir(parents=True, exist_ok=True)


def load_rows(source: Path) -> tuple[str, list[str], list[dict[str, Any]]]:
    workbook = load_workbook(source, read_only=True, data_only=True)
    worksheet = workbook.active
    headers = [normalized_scalar(cell.value) for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    rows: list[dict[str, Any]] = []

    for excel_row, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        row = {header: value for header, value in zip(headers, values)}
        row["__source_excel_row"] = excel_row
        rows.append(row)

    return worksheet.title, headers, rows


def build_manifests(
    source: Path,
    output: Path,
    limit: int | None = None,
) -> tuple[list[dict[str, str]], list[MediaItem]]:
    sheet_name, headers, rows = load_rows(source)
    if limit is not None:
        rows = rows[:limit]

    records: list[dict[str, str]] = []
    media_items: list[MediaItem] = []

    for row_position, row in enumerate(rows, start=1):
        excel_row = int(row["__source_excel_row"])
        patient_id = first_available(
            row,
            ["\u60a3\u8005id", "\u8001\u6765\u53f7"],
            f"row{excel_row:04d}",
        )
        collection_no = first_available(
            row,
            ["\u91c7\u96c6\u7f16\u53f7"],
            f"row{excel_row:04d}",
        )
        record_id = f"row{excel_row:04d}_pid{patient_id}_collect{collection_no}"

        record: dict[str, str] = {
            "record_id": record_id,
            "source_excel_row": str(excel_row),
            "source_sheet": sheet_name,
            "source_file": source.name,
        }
        for header in headers:
            record[header] = normalized_scalar(row.get(header))

        image_count = 0
        video_count = 0
        media_count = 0

        for header in headers:
            role = MEDIA_COLUMN_ROLES.get(header)
            if role is None:
                continue

            urls = extract_urls(row.get(header))
            local_paths: list[str] = []

            for index_in_field, url in enumerate(urls, start=1):
                media_type = media_type_from_url(url)
                ext = extension_from_url(url, media_type)
                media_count += 1
                if media_type == "image":
                    image_count += 1
                    type_dir = "images"
                elif media_type == "video":
                    video_count += 1
                    type_dir = "videos"
                else:
                    type_dir = "unknown"

                media_id = f"{record_id}__{role}_{index_in_field:02d}"
                filename = f"{media_id}{ext}"
                local_path = output / "media" / type_dir / record_id / filename
                local_paths.append(local_path.relative_to(output).as_posix())

                media_items.append(
                    MediaItem(
                        media_id=media_id,
                        record_id=record_id,
                        source_excel_row=excel_row,
                        source_sheet=sheet_name,
                        patient_id=patient_id,
                        collection_no=collection_no,
                        field_name=header,
                        media_role=role,
                        media_index_in_field=index_in_field,
                        media_type=media_type,
                        url=url,
                        local_path=local_path,
                        filename=filename,
                    )
                )

            record[f"{header}__local_paths"] = "|".join(local_paths)

        record["media_count"] = str(media_count)
        record["image_count"] = str(image_count)
        record["video_count"] = str(video_count)
        records.append(record)

    return records, media_items


def download_one(
    item: MediaItem,
    output: Path,
    timeout: float,
    overwrite: bool,
    retries: int,
) -> dict[str, str]:
    relative_path = item.local_path.relative_to(output).as_posix()
    result = {
        "media_id": item.media_id,
        "record_id": item.record_id,
        "source_excel_row": str(item.source_excel_row),
        "source_sheet": item.source_sheet,
        "patient_id": item.patient_id,
        "collection_no": item.collection_no,
        "field_name": item.field_name,
        "media_role": item.media_role,
        "media_index_in_field": str(item.media_index_in_field),
        "media_type": item.media_type,
        "url": item.url,
        "local_path": relative_path,
        "filename": item.filename,
        "download_status": "",
        "http_status": "",
        "content_type": "",
        "bytes": "",
        "sha256": "",
        "error": "",
    }

    item.local_path.parent.mkdir(parents=True, exist_ok=True)

    if item.local_path.exists() and item.local_path.stat().st_size > 0 and not overwrite:
        result["download_status"] = "exists"
        result["bytes"] = str(item.local_path.stat().st_size)
        result["sha256"] = sha256_path(item.local_path)
        return result

    request = urllib.request.Request(
        item.url,
        headers={
            "User-Agent": "Mozilla/5.0 FaceSymAi dataset builder",
            "Accept": "*/*",
        },
    )

    last_error = ""
    for attempt in range(1, retries + 2):
        tmp_name = ""
        try:
            with urllib.request.urlopen(request, timeout=timeout) as response:
                result["http_status"] = str(getattr(response, "status", ""))
                result["content_type"] = response.headers.get("Content-Type", "")
                fd, tmp_name = tempfile.mkstemp(
                    prefix=f".{item.filename}.",
                    suffix=".download",
                    dir=str(item.local_path.parent),
                )
                with os.fdopen(fd, "wb") as tmp_file:
                    shutil.copyfileobj(response, tmp_file, length=1024 * 1024)
            tmp_path = Path(tmp_name)
            if tmp_path.stat().st_size == 0:
                raise RuntimeError("downloaded file is empty")
            tmp_path.replace(item.local_path)
            result["download_status"] = "downloaded"
            result["bytes"] = str(item.local_path.stat().st_size)
            result["sha256"] = sha256_path(item.local_path)
            result["error"] = ""
            return result
        except urllib.error.HTTPError as exc:
            last_error = f"HTTPError {exc.code}: {exc.reason}"
            result["http_status"] = str(exc.code)
        except Exception as exc:  # noqa: BLE001 - preserve failure in manifest
            last_error = f"{type(exc).__name__}: {exc}"
        finally:
            if tmp_name:
                tmp_path = Path(tmp_name)
                if tmp_path.exists():
                    tmp_path.unlink()

        if attempt <= retries:
            time.sleep(min(2.0 * attempt, 8.0))

    result["download_status"] = "failed"
    result["error"] = last_error
    return result


def write_csv(path: Path, rows: Iterable[dict[str, str]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_records_jsonl(
    path: Path,
    records: list[dict[str, str]],
    media_results: list[dict[str, str]],
) -> None:
    media_by_record: dict[str, list[dict[str, str]]] = {}
    for item in media_results:
        media_by_record.setdefault(item["record_id"], []).append(item)

    with path.open("w", encoding="utf-8") as handle:
        for record in records:
            payload = {
                "record": record,
                "media": media_by_record.get(record["record_id"], []),
            }
            handle.write(json.dumps(payload, ensure_ascii=False) + "\n")


def write_readme(
    output: Path,
    source: Path,
    records: list[dict[str, str]],
    media_results: list[dict[str, str]],
    source_sha256: str,
) -> None:
    status_counts: dict[str, int] = {}
    type_counts: dict[str, int] = {}
    for item in media_results:
        status_counts[item["download_status"]] = status_counts.get(item["download_status"], 0) + 1
        type_counts[item["media_type"]] = type_counts.get(item["media_type"], 0) + 1

    lines = [
        "# Stroke Media Dataset",
        "",
        f"- Source workbook: `{source.name}`",
        f"- Source SHA256: `{source_sha256}`",
        f"- Generated at: `{datetime.now().isoformat(timespec='seconds')}`",
        f"- Records: `{len(records)}`",
        f"- Media URLs: `{len(media_results)}`",
        f"- Media types: `{json.dumps(type_counts, ensure_ascii=False, sort_keys=True)}`",
        f"- Download status: `{json.dumps(status_counts, ensure_ascii=False, sort_keys=True)}`",
        "",
        "## Layout",
        "",
        "- `metadata/records.csv`: one row per Excel record, with original fields and local media path columns.",
        "- `metadata/media_manifest.csv`: one row per media URL/file, preserving the original URL and local path.",
        "- `metadata/records.jsonl`: one JSON object per record, grouping its media files.",
        "- `media/images/<record_id>/`: downloaded images.",
        "- `media/videos/<record_id>/`: downloaded videos.",
        "- `media/unknown/<record_id>/`: downloaded files whose type was not inferred from the URL.",
        "",
        "## Notes",
        "",
        "- File and directory names intentionally avoid patient names.",
        "- The metadata still preserves source workbook fields, including sensitive medical/person data.",
        "- Re-run the script to resume failed or interrupted downloads; existing non-empty files are reused by default.",
    ]
    (output / "README.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE, help="Source xlsx workbook.")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="Output dataset directory.")
    parser.add_argument("--workers", type=int, default=8, help="Concurrent download workers.")
    parser.add_argument("--timeout", type=float, default=45.0, help="Download timeout in seconds.")
    parser.add_argument("--retries", type=int, default=2, help="Retries per media URL after the first attempt.")
    parser.add_argument("--limit", type=int, default=None, help="Only process the first N records.")
    parser.add_argument("--no-download", action="store_true", help="Only generate metadata manifests.")
    parser.add_argument("--overwrite", action="store_true", help="Re-download files that already exist.")
    return parser.parse_args(argv)


def main(argv: list[str]) -> int:
    args = parse_args(argv)
    source = args.source.resolve()
    output = args.output.resolve()

    if not source.exists():
        print(f"Source workbook not found: {source}", file=sys.stderr)
        return 2
    if args.workers < 1:
        print("--workers must be >= 1", file=sys.stderr)
        return 2

    prepare_dirs(output)
    source_sha = sha256_file(source)
    records, media_items = build_manifests(source, output, limit=args.limit)

    record_fields = BASE_RECORD_FIELDS[:]
    if records:
        for key in records[0]:
            if key not in record_fields:
                record_fields.append(key)

    if args.no_download:
        media_results = []
        for item in media_items:
            media_results.append(
                {
                    "media_id": item.media_id,
                    "record_id": item.record_id,
                    "source_excel_row": str(item.source_excel_row),
                    "source_sheet": item.source_sheet,
                    "patient_id": item.patient_id,
                    "collection_no": item.collection_no,
                    "field_name": item.field_name,
                    "media_role": item.media_role,
                    "media_index_in_field": str(item.media_index_in_field),
                    "media_type": item.media_type,
                    "url": item.url,
                    "local_path": item.local_path.relative_to(output).as_posix(),
                    "filename": item.filename,
                    "download_status": "not_downloaded",
                    "http_status": "",
                    "content_type": "",
                    "bytes": "",
                    "sha256": "",
                    "error": "",
                }
            )
    else:
        media_results = []
        total = len(media_items)
        completed = 0
        with concurrent.futures.ThreadPoolExecutor(max_workers=args.workers) as executor:
            futures = [
                executor.submit(
                    download_one,
                    item,
                    output,
                    args.timeout,
                    args.overwrite,
                    args.retries,
                )
                for item in media_items
            ]
            for future in concurrent.futures.as_completed(futures):
                media_results.append(future.result())
                completed += 1
                if completed % 100 == 0 or completed == total:
                    print(f"download progress: {completed}/{total}", flush=True)

        media_results.sort(
            key=lambda row: (
                int(row["source_excel_row"]),
                row["record_id"],
                row["media_role"],
                int(row["media_index_in_field"]),
            )
        )

    write_csv(output / "metadata" / "records.csv", records, record_fields)
    write_csv(output / "metadata" / "media_manifest.csv", media_results, MEDIA_MANIFEST_FIELDS)
    write_records_jsonl(output / "metadata" / "records.jsonl", records, media_results)
    write_readme(output, source, records, media_results, source_sha)

    failed = sum(1 for item in media_results if item["download_status"] == "failed")
    downloaded = sum(1 for item in media_results if item["download_status"] == "downloaded")
    existing = sum(1 for item in media_results if item["download_status"] == "exists")
    print(
        json.dumps(
            {
                "output": str(output),
                "records": len(records),
                "media_urls": len(media_results),
                "downloaded": downloaded,
                "existing": existing,
                "failed": failed,
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 1 if failed else 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
