#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import os
import re
import shutil
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Mapping
from xml.etree import ElementTree


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE = PROJECT_ROOT / "脑卒中预警报告老来健康app线上_2026-05-08.xlsx"
DEFAULT_MEDIA_DATASET = PROJECT_ROOT / "datasets" / "stroke_warning_app_media_dataset_20260508"
DEFAULT_OUTPUT = PROJECT_ROOT / "datasets" / "stroke_warning_app_rule_test_set_20260508"

PATIENT_ID_FIELD = "老来号"
PHONE_FIELD = "手机号"
RISK_FIELD = "风险等级"
BASE_DISEASE_FIELD = "您有没有以下这些基础小毛病？(可多选)"
SMOKE_FIELD = "您平时有抽烟的习惯吗？"
EXERCISE_FIELD = "您经常运动吗？(每周大于等于2次)"
OVERWEIGHT_FIELD = "您目前是否明显超重或肥胖？"
WEAKNESS_FIELD = "最近有没有突然觉得手脚没力气，拿不住东西？"
NUMBNESS_FIELD = "最近有没有突然觉得半边脸，或者半边身子一阵发麻？"
PRIOR_STROKE_FIELD = "您以前有没有得过中风（比如脑梗、脑出血），或者是“小中风”？"
FAMILY_STROKE_FIELD = "您的家里人 (父母或兄弟姐妹），有人得过脑卒中吗？"

NORMAL_VALUES = {"", "无", "否", "没有"}

SELECTED_RECORD_FIELDS = [
    "record_id",
    "patient_sample_id",
    "patient_id",
    "source_excel_row",
    "label_group",
    "label_binary",
    "selection_status",
    "rule_reasons",
    "risk_level",
    "sex",
    "age",
    "evaluation_time",
    "base_disease",
    "smoking",
    "exercise",
    "overweight",
    "sudden_weakness",
    "sudden_numbness",
    "prior_stroke",
    "family_stroke",
]

PATIENT_FIELDS = [
    "patient_sample_id",
    "patient_id",
    "label_group",
    "label_binary",
    "record_count",
    "selected_record_count",
    "positive_record_count",
    "normal_record_count",
    "media_count",
    "image_count",
    "video_count",
    "source_excel_rows",
    "selected_record_ids",
    "risk_levels",
    "rule_reasons",
    "sex_values",
    "age_values",
    "patient_dir",
]

MEDIA_FIELDS = [
    "media_id",
    "patient_sample_id",
    "patient_id",
    "label_group",
    "label_binary",
    "record_id",
    "source_excel_row",
    "field_name",
    "media_role",
    "media_type",
    "source_media_path",
    "organized_path",
    "link_mode",
    "bytes",
    "sha256",
]

EXCLUDED_RECORD_FIELDS = [
    "record_id",
    "patient_sample_id",
    "patient_id",
    "source_excel_row",
    "risk_level",
    "exclude_reason",
    "rule_reasons",
]


@dataclass
class RuleRecord:
    record_id: str
    source_excel_row: int
    patient_id: str
    row: dict[str, str]
    positive_reasons: list[str]
    is_normal_negative: bool


@dataclass
class PatientDecision:
    patient_id: str
    patient_sample_id: str
    label_group: str
    label_binary: str
    records: list[RuleRecord] = field(default_factory=list)
    selected_records: list[RuleRecord] = field(default_factory=list)


def main() -> None:
    args = parse_args()
    source = args.source.resolve()
    media_dataset = args.media_dataset.resolve()
    output = args.output.resolve()

    workbook_records = load_rule_records(source)
    decisions = decide_patients(workbook_records)
    media_manifest = read_csv(media_dataset / "metadata" / "media_manifest.csv")

    record_rows, excluded_rows = build_record_rows(decisions)
    prepare_output(output, media_dataset)
    media_rows = build_media_rows(decisions, media_manifest, media_dataset, output, args.link_mode)
    patient_rows = build_patient_rows(decisions, media_rows, output)
    summary = build_summary(source, media_dataset, output, workbook_records, decisions, record_rows, excluded_rows, media_rows)

    write_csv(output / "metadata" / "rule_labeled_records.csv", record_rows, SELECTED_RECORD_FIELDS)
    write_csv(output / "metadata" / "patient_samples.csv", patient_rows, PATIENT_FIELDS)
    write_csv(output / "metadata" / "media_index.csv", media_rows, MEDIA_FIELDS)
    write_csv(output / "metadata" / "excluded_records.csv", excluded_rows, EXCLUDED_RECORD_FIELDS)
    write_json(output / "metadata" / "summary.json", summary)
    write_jsonl(output / "metadata" / "rule_labeled_records.jsonl", record_rows)
    write_report(output / "reports" / "01_rule_test_set.md", summary, patient_rows)
    write_readme(output / "README.md", summary)

    print(f"Wrote {output / 'metadata' / 'rule_labeled_records.csv'}")
    print(f"Wrote {output / 'metadata' / 'patient_samples.csv'}")
    print(f"Wrote {output / 'metadata' / 'media_index.csv'}")
    print(f"Wrote {output / 'metadata' / 'excluded_records.csv'}")
    print(f"Wrote {output / 'metadata' / 'summary.json'}")
    print(f"Wrote {output / 'reports' / '01_rule_test_set.md'}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build a rule-labeled test set from the stroke warning app workbook.")
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE, help="Source app warning workbook.")
    parser.add_argument("--media-dataset", type=Path, default=DEFAULT_MEDIA_DATASET, help="Downloaded app media dataset root.")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="Output rule test set root.")
    parser.add_argument("--link-mode", choices=("hardlink", "copy", "symlink"), default="hardlink", help="How to place media files.")
    return parser.parse_args()


def prepare_output(output: Path, media_dataset: Path) -> None:
    if output == PROJECT_ROOT:
        raise ValueError("Output directory cannot be the project root.")
    if output == media_dataset:
        raise ValueError("Output directory cannot be the source media dataset.")
    for name in ("患病", "不患病", "metadata", "reports"):
        path = output / name
        if path.exists():
            shutil.rmtree(path)
    readme = output / "README.md"
    if readme.exists():
        readme.unlink()


def normalized_scalar(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def load_rule_records(source: Path) -> list[RuleRecord]:
    raw_rows = load_workbook_rows(source)
    if not raw_rows:
        return []
    headers = [normalized_scalar(value) for value in raw_rows[0]]
    rows: list[RuleRecord] = []
    for excel_row, values in enumerate(raw_rows[1:], start=2):
        row = {header: normalized_scalar(value) for header, value in zip(headers, values)}
        patient_id = row.get(PATIENT_ID_FIELD, "") or f"row{excel_row:04d}"
        rows.append(
            RuleRecord(
                record_id=record_id_for(excel_row, patient_id),
                source_excel_row=excel_row,
                patient_id=patient_id,
                row=row,
                positive_reasons=positive_reasons(row),
                is_normal_negative=is_all_normal_low_risk(row),
            )
        )
    return rows


def load_workbook_rows(source: Path) -> list[list[Any]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        return load_xlsx_rows_stdlib(source)
    workbook = load_workbook(source, read_only=True, data_only=True)
    worksheet = workbook.active
    return [list(row) for row in worksheet.iter_rows(values_only=True)]


def load_xlsx_rows_stdlib(source: Path) -> list[list[str]]:
    with zipfile.ZipFile(source) as archive:
        shared_strings = load_shared_strings(archive)
        sheet_name = first_sheet_path(archive)
        root = ElementTree.fromstring(archive.read(sheet_name))

    rows: list[list[str]] = []
    ns = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    for row_element in root.findall(".//x:sheetData/x:row", ns):
        values_by_index: dict[int, str] = {}
        for cell in row_element.findall("x:c", ns):
            ref = cell.attrib.get("r", "")
            column_index = column_index_from_ref(ref)
            values_by_index[column_index] = cell_value(cell, shared_strings, ns)
        if values_by_index:
            width = max(values_by_index) + 1
            rows.append([values_by_index.get(index, "") for index in range(width)])
    return rows


def load_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    try:
        payload = archive.read("xl/sharedStrings.xml")
    except KeyError:
        return []
    root = ElementTree.fromstring(payload)
    ns = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    values: list[str] = []
    for item in root.findall("x:si", ns):
        texts = [node.text or "" for node in item.findall(".//x:t", ns)]
        values.append("".join(texts))
    return values


def first_sheet_path(archive: zipfile.ZipFile) -> str:
    if "xl/worksheets/sheet1.xml" in archive.namelist():
        return "xl/worksheets/sheet1.xml"
    candidates = sorted(name for name in archive.namelist() if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"))
    if not candidates:
        raise FileNotFoundError("No worksheet XML found in xlsx file.")
    return candidates[0]


def column_index_from_ref(ref: str) -> int:
    letters = re.sub(r"[^A-Z]", "", ref.upper())
    result = 0
    for char in letters:
        result = result * 26 + (ord(char) - ord("A") + 1)
    return max(result - 1, 0)


def cell_value(cell: ElementTree.Element, shared_strings: list[str], ns: Mapping[str, str]) -> str:
    cell_type = cell.attrib.get("t", "")
    if cell_type == "inlineStr":
        texts = [node.text or "" for node in cell.findall(".//x:t", ns)]
        return "".join(texts)
    value_node = cell.find("x:v", ns)
    if value_node is None or value_node.text is None:
        return ""
    raw = value_node.text
    if cell_type == "s":
        index = int(raw)
        return shared_strings[index] if 0 <= index < len(shared_strings) else ""
    return raw


def record_id_for(excel_row: int, patient_id: str) -> str:
    return f"row{excel_row:04d}_pid{clean_id(patient_id)}_collectrow{excel_row:04d}"


def clean_id(value: str) -> str:
    cleaned = re.sub(r"\.0$", "", str(value).strip())
    cleaned = re.sub(r"[^A-Za-z0-9_-]+", "-", cleaned).strip("-")
    return cleaned or "unknown"


def clean_path_part(value: str, fallback: str) -> str:
    text = str(value or "").strip() or fallback
    text = re.sub(r"[\\/:*?\"<>|]+", "_", text)
    text = re.sub(r"\s+", "_", text).strip("._ ")
    return text or fallback


def positive_reasons(row: Mapping[str, str]) -> list[str]:
    has_emergency_risk = row.get(RISK_FIELD, "") == "紧急风险"
    has_prior_stroke = is_positive_answer(row.get(PRIOR_STROKE_FIELD, ""))
    has_family_stroke = is_positive_answer(row.get(FAMILY_STROKE_FIELD, ""))
    if has_emergency_risk and has_prior_stroke and has_family_stroke:
        return ["risk_level_emergency", "prior_stroke_yes", "family_stroke_yes"]
    return []


def is_positive_answer(value: str) -> bool:
    return normalized_scalar(value) not in NORMAL_VALUES


def is_all_normal_low_risk(row: Mapping[str, str]) -> bool:
    return (
        row.get(RISK_FIELD, "") == "低风险"
        and row.get(BASE_DISEASE_FIELD, "") == "无"
        and row.get(SMOKE_FIELD, "") == "无"
        and row.get(EXERCISE_FIELD, "") == "是"
        and row.get(OVERWEIGHT_FIELD, "") == "否"
        and row.get(WEAKNESS_FIELD, "") == "无"
        and row.get(NUMBNESS_FIELD, "") == "无"
        and row.get(PRIOR_STROKE_FIELD, "") == "否"
        and row.get(FAMILY_STROKE_FIELD, "") == "无"
    )


def decide_patients(records: list[RuleRecord]) -> dict[str, PatientDecision]:
    grouped: dict[str, list[RuleRecord]] = defaultdict(list)
    for record in records:
        grouped[record.patient_id].append(record)

    decisions: dict[str, PatientDecision] = {}
    for patient_id, patient_records in sorted(grouped.items()):
        patient_sample_id = f"pid{clean_id(patient_id)}"
        positive_records = [record for record in patient_records if record.positive_reasons]
        normal_records = [record for record in patient_records if record.is_normal_negative]
        if positive_records:
            decision = PatientDecision(
                patient_id=patient_id,
                patient_sample_id=patient_sample_id,
                label_group="患病",
                label_binary="1",
                records=patient_records,
                selected_records=positive_records,
            )
        elif normal_records:
            decision = PatientDecision(
                patient_id=patient_id,
                patient_sample_id=patient_sample_id,
                label_group="不患病",
                label_binary="0",
                records=patient_records,
                selected_records=normal_records,
            )
        else:
            decision = PatientDecision(
                patient_id=patient_id,
                patient_sample_id=patient_sample_id,
                label_group="未纳入",
                label_binary="",
                records=patient_records,
                selected_records=[],
            )
        decisions[patient_id] = decision
    return decisions


def build_record_rows(decisions: Mapping[str, PatientDecision]) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    selected_rows: list[dict[str, str]] = []
    excluded_rows: list[dict[str, str]] = []
    selected_record_ids = {record.record_id for decision in decisions.values() for record in decision.selected_records}
    for decision in decisions.values():
        for record in decision.records:
            reasons = record.positive_reasons[:]
            if record.is_normal_negative:
                reasons.append("low_risk_all_indicators_normal")
            if record.record_id in selected_record_ids:
                selected_rows.append(record_row(decision, record, "selected", reasons))
            else:
                excluded_rows.append(
                    {
                        "record_id": record.record_id,
                        "patient_sample_id": decision.patient_sample_id,
                        "patient_id": decision.patient_id,
                        "source_excel_row": str(record.source_excel_row),
                        "risk_level": record.row.get(RISK_FIELD, ""),
                        "exclude_reason": exclude_reason(decision, record),
                        "rule_reasons": ";".join(reasons),
                    }
                )
    return selected_rows, excluded_rows


def record_row(decision: PatientDecision, record: RuleRecord, selection_status: str, reasons: list[str]) -> dict[str, str]:
    row = record.row
    return {
        "record_id": record.record_id,
        "patient_sample_id": decision.patient_sample_id,
        "patient_id": decision.patient_id,
        "source_excel_row": str(record.source_excel_row),
        "label_group": decision.label_group,
        "label_binary": decision.label_binary,
        "selection_status": selection_status,
        "rule_reasons": ";".join(reasons),
        "risk_level": row.get(RISK_FIELD, ""),
        "sex": row.get("性别", ""),
        "age": row.get("年龄", ""),
        "evaluation_time": row.get("评估时间", ""),
        "base_disease": row.get(BASE_DISEASE_FIELD, ""),
        "smoking": row.get(SMOKE_FIELD, ""),
        "exercise": row.get(EXERCISE_FIELD, ""),
        "overweight": row.get(OVERWEIGHT_FIELD, ""),
        "sudden_weakness": row.get(WEAKNESS_FIELD, ""),
        "sudden_numbness": row.get(NUMBNESS_FIELD, ""),
        "prior_stroke": row.get(PRIOR_STROKE_FIELD, ""),
        "family_stroke": row.get(FAMILY_STROKE_FIELD, ""),
    }


def exclude_reason(decision: PatientDecision, record: RuleRecord) -> str:
    if decision.label_group == "未纳入":
        return "no_rule_match"
    if decision.label_group == "患病" and record.is_normal_negative:
        return "patient_positive_precedence_over_normal_record"
    return "non_qualifying_record_for_selected_patient"


def build_media_rows(
    decisions: Mapping[str, PatientDecision],
    media_manifest: list[Mapping[str, str]],
    media_dataset: Path,
    output: Path,
    link_mode: str,
) -> list[dict[str, str]]:
    selected_record_ids = {record.record_id for decision in decisions.values() for record in decision.selected_records}
    decision_by_record = {
        record.record_id: decision
        for decision in decisions.values()
        for record in decision.selected_records
    }
    record_by_id = {
        record.record_id: record
        for decision in decisions.values()
        for record in decision.selected_records
    }
    output_rows: list[dict[str, str]] = []
    for media in media_manifest:
        record_id = media.get("record_id", "")
        if record_id not in selected_record_ids:
            continue
        if media.get("download_status", "") not in {"downloaded", "exists"}:
            continue
        source_media = media_dataset / media.get("local_path", "")
        if not source_media.exists():
            continue
        decision = decision_by_record[record_id]
        rule_record = record_by_id[record_id]
        media_type = media.get("media_type", "")
        bucket = "images" if media_type == "image" else "videos" if media_type == "video" else "unknown"
        label_dir = clean_path_part(decision.label_group, "unlabeled")
        patient_dir = clean_path_part(decision.patient_sample_id, "patient")
        filename = media.get("filename") or source_media.name
        target = output / label_dir / patient_dir / bucket / filename
        actual_mode = link_media(source_media, target, link_mode)
        output_rows.append(
            {
                "media_id": media.get("media_id", ""),
                "patient_sample_id": decision.patient_sample_id,
                "patient_id": decision.patient_id,
                "label_group": decision.label_group,
                "label_binary": decision.label_binary,
                "record_id": record_id,
                "source_excel_row": str(rule_record.source_excel_row),
                "field_name": media.get("field_name", ""),
                "media_role": media.get("media_role", ""),
                "media_type": media_type,
                "source_media_path": source_media.as_posix(),
                "organized_path": target.relative_to(output).as_posix(),
                "link_mode": actual_mode,
                "bytes": media.get("bytes", ""),
                "sha256": media.get("sha256", ""),
            }
        )
    return sorted(output_rows, key=lambda item: (item["label_group"], item["patient_sample_id"], item["record_id"], item["media_role"], item["media_id"]))


def link_media(source: Path, target: Path, mode: str) -> str:
    target.parent.mkdir(parents=True, exist_ok=True)
    if target.exists() and target.stat().st_size == source.stat().st_size:
        return "exists"
    if target.exists():
        target.unlink()
    if mode == "copy":
        shutil.copy2(source, target)
        return "copy"
    if mode == "symlink":
        target.symlink_to(source)
        return "symlink"
    try:
        os.link(source, target)
        return "hardlink"
    except OSError:
        shutil.copy2(source, target)
        return "copy"


def build_patient_rows(
    decisions: Mapping[str, PatientDecision],
    media_rows: list[Mapping[str, str]],
    output: Path,
) -> list[dict[str, str]]:
    media_by_patient: dict[str, list[Mapping[str, str]]] = defaultdict(list)
    for row in media_rows:
        media_by_patient[row["patient_sample_id"]].append(row)

    rows: list[dict[str, str]] = []
    for decision in decisions.values():
        if decision.label_group == "未纳入":
            continue
        patient_media = media_by_patient.get(decision.patient_sample_id, [])
        positive_count = sum(1 for record in decision.records if record.positive_reasons)
        normal_count = sum(1 for record in decision.records if record.is_normal_negative)
        risk_levels = sorted({record.row.get(RISK_FIELD, "") for record in decision.records if record.row.get(RISK_FIELD, "")})
        rule_reasons = sorted(
            {
                reason
                for record in decision.selected_records
                for reason in (record.positive_reasons or ["low_risk_all_indicators_normal"])
            }
        )
        rows.append(
            {
                "patient_sample_id": decision.patient_sample_id,
                "patient_id": decision.patient_id,
                "label_group": decision.label_group,
                "label_binary": decision.label_binary,
                "record_count": str(len(decision.records)),
                "selected_record_count": str(len(decision.selected_records)),
                "positive_record_count": str(positive_count),
                "normal_record_count": str(normal_count),
                "media_count": str(len(patient_media)),
                "image_count": str(sum(1 for row in patient_media if row.get("media_type") == "image")),
                "video_count": str(sum(1 for row in patient_media if row.get("media_type") == "video")),
                "source_excel_rows": ";".join(str(record.source_excel_row) for record in decision.selected_records),
                "selected_record_ids": ";".join(record.record_id for record in decision.selected_records),
                "risk_levels": ";".join(risk_levels),
                "rule_reasons": ";".join(rule_reasons),
                "sex_values": ";".join(sorted({record.row.get("性别", "") for record in decision.records if record.row.get("性别", "")})),
                "age_values": ";".join(sorted({record.row.get("年龄", "") for record in decision.records if record.row.get("年龄", "")})),
                "patient_dir": (Path(clean_path_part(decision.label_group, "unlabeled")) / clean_path_part(decision.patient_sample_id, "patient")).as_posix(),
            }
        )
    return sorted(rows, key=lambda item: (item["label_group"], item["patient_sample_id"]))


def build_summary(
    source: Path,
    media_dataset: Path,
    output: Path,
    all_records: list[RuleRecord],
    decisions: Mapping[str, PatientDecision],
    selected_rows: list[Mapping[str, str]],
    excluded_rows: list[Mapping[str, str]],
    media_rows: list[Mapping[str, str]],
) -> dict[str, Any]:
    selected_decisions = [decision for decision in decisions.values() if decision.label_group != "未纳入"]
    return {
        "source_workbook": source.as_posix(),
        "source_media_dataset": media_dataset.as_posix(),
        "output_dataset": output.as_posix(),
        "label_policy": {
            "positive": "patient is 患病 if any single record simultaneously has 风险等级=紧急风险 AND prior stroke answer is positive AND family stroke answer is positive",
            "negative": "patient is 不患病 if no positive record exists and at least one record is 低风险 with all questionnaire indicators normal",
            "precedence": "positive wins over low-risk normal records for the same patient",
        },
        "source_record_count": len(all_records),
        "source_patient_count": len(decisions),
        "selected_patient_count": len(selected_decisions),
        "selected_record_count": len(selected_rows),
        "excluded_record_count": len(excluded_rows),
        "media_count": len(media_rows),
        "image_count": sum(1 for row in media_rows if row.get("media_type") == "image"),
        "video_count": sum(1 for row in media_rows if row.get("media_type") == "video"),
        "patients_by_label": dict(sorted(Counter(decision.label_group for decision in selected_decisions).items())),
        "records_by_label": dict(sorted(Counter(row["label_group"] for row in selected_rows).items())),
        "media_by_label": dict(sorted(Counter(row["label_group"] for row in media_rows).items())),
        "media_by_role": dict(sorted(Counter(row["media_role"] for row in media_rows).items())),
        "excluded_patients": sum(1 for decision in decisions.values() if decision.label_group == "未纳入"),
        "positive_precedence_patient_count": sum(
            1
            for decision in decisions.values()
            if decision.label_group == "患病" and any(record.is_normal_negative for record in decision.records)
        ),
    }


def write_csv(path: Path, rows: Iterable[Mapping[str, str]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_json(path: Path, payload: Mapping[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def write_jsonl(path: Path, rows: Iterable[Mapping[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n")


def write_report(path: Path, summary: Mapping[str, Any], patient_rows: list[Mapping[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "# 01 脑卒中预警 App 规则测试集",
        "",
        f"源 Excel：`{summary['source_workbook']}`",
        f"源媒体数据集：`{summary['source_media_dataset']}`",
        "",
        "## 标签规则",
        "",
        "- 患病：同一条记录同时满足 `风险等级=紧急风险`、`曾经得过中风=是`、`家人得过脑卒中=有`。",
        "- 不患病：该患者没有任何阳性记录，且至少一条记录为 `低风险`，同时基础病、抽烟、超重、突发无力、突发麻木、中风史、家族史均正常，且经常运动为 `是`。",
        "- 同一患者同时存在阳性和低风险全正常记录时，阳性优先。",
        "",
        "## 汇总",
        "",
        f"- 源记录数：`{summary['source_record_count']}`",
        f"- 源患者数：`{summary['source_patient_count']}`",
        f"- 纳入患者数：`{summary['selected_patient_count']}`",
        f"- 纳入记录数：`{summary['selected_record_count']}`",
        f"- 排除记录数：`{summary['excluded_record_count']}`",
        f"- 媒体数：`{summary['media_count']}`，图片 `{summary['image_count']}`，视频 `{summary['video_count']}`",
        f"- 患者标签分布：`{json.dumps(summary['patients_by_label'], ensure_ascii=False, sort_keys=True)}`",
        f"- 媒体 role 分布：`{json.dumps(summary['media_by_role'], ensure_ascii=False, sort_keys=True)}`",
        f"- 阳性优先处理患者数：`{summary['positive_precedence_patient_count']}`",
        "",
        "## 产物",
        "",
        "- 患者级标签：`metadata/patient_samples.csv`",
        "- 记录级标签：`metadata/rule_labeled_records.csv`",
        "- 媒体索引：`metadata/media_index.csv`",
        "- 排除记录：`metadata/excluded_records.csv`",
        "- 汇总 JSON：`metadata/summary.json`",
        "",
        "## 患者样本预览",
        "",
    ]
    lines.extend(
        markdown_table(
            ["patient", "label", "records", "media", "rows", "reasons"],
            [
                [
                    row["patient_sample_id"],
                    row["label_group"],
                    row["selected_record_count"],
                    row["media_count"],
                    row["source_excel_rows"],
                    row["rule_reasons"],
                ]
                for row in patient_rows[:80]
            ],
        )
    )
    path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


def write_readme(path: Path, summary: Mapping[str, Any]) -> None:
    lines = [
        "# Stroke Warning App Rule Test Set",
        "",
        "该数据集由 `脑卒中预警报告老来健康app线上_2026-05-08.xlsx` 按规则生成。",
        "",
        f"- 患者标签分布：`{json.dumps(summary['patients_by_label'], ensure_ascii=False, sort_keys=True)}`",
        f"- 媒体数：`{summary['media_count']}`",
        "- 详细说明见 `reports/01_rule_test_set.md`。",
        "",
    ]
    path.write_text("\n".join(lines), encoding="utf-8")


def markdown_table(headers: list[str], rows: list[list[Any]]) -> list[str]:
    output = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    for row in rows:
        output.append("| " + " | ".join(str(value) for value in row) + " |")
    return output


if __name__ == "__main__":
    main()
